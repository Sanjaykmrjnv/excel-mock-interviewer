import openpyxl
from typing import Dict, Any

def evaluate_submission(filepath: str) -> Dict[str, Any]:
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Read header row
    headers = [ (cell.value or "").strip() if isinstance(cell.value, str) else cell.value for cell in ws[1] ]
    # Try find columns by header name
    def find_col(name, fallback_idx):
        try:
            return headers.index(name)  # 0-based
        except ValueError:
            return fallback_idx

    idx_qty = find_col("Qty", 0)
    idx_price = find_col("Price", 1)
    idx_total = find_col("Total", 2)

    row_mismatches = []
    totals_calc = []
    row_count = 0

    for r in range(2, ws.max_row + 1):
        qty = ws.cell(row=r, column=idx_qty+1).value
        price = ws.cell(row=r, column=idx_price+1).value
        total = ws.cell(row=r, column=idx_total+1).value
        if qty is None and price is None and total is None:
            continue
        row_count += 1
        try:
            q = float(qty or 0)
            p = float(price or 0)
            expected = q * p
        except Exception:
            row_mismatches.append({
                "row": r,
                "reason": "non-numeric qty/price",
                "qty": qty, "price": price, "total": total
            })
            totals_calc.append(0)
            continue

        totals_calc.append(expected)
        # compare numeric with tolerance
        if total is None or abs((float(total or 0) - expected)) > 1e-6:
            row_mismatches.append({
                "row": r,
                "expected": expected,
                "actual": total
            })

    # G1 check
    g1_actual = None
    try:
        g1_actual = ws["G1"].value
    except Exception:
        pass

    expected_sum = sum(totals_calc)
    g1_mismatch = None
    try:
        if g1_actual is None or abs(float(g1_actual or 0) - expected_sum) > 1e-6:
            g1_mismatch = {"expected": expected_sum, "actual": g1_actual}
    except Exception:
        g1_mismatch = {"expected": expected_sum, "actual": g1_actual}

    passed = (len(row_mismatches) == 0) and (g1_mismatch is None)

    return {
        "row_count": row_count,
        "row_mismatches": row_mismatches,
        "g1_mismatch": g1_mismatch,
        "expected_sum": expected_sum,
        "g1_actual": g1_actual,
        "pass": passed
    }
